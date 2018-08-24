from bs4 import BeautifulSoup
import requests
import re
import openpyxl

### INPUT TEST URLS HERE ###

urlstring = ["lazada.com.my","maybank2u.com.my","airasia.com","shopee.com.my","cimbclicks.com.my","lowyat.net","paultan.org","malaysiakini.com","thestar.com.my","mudah.my","pbebank.com","mkyong.com","hide.me","moe.gov.my","malaysiaairlines.com","11street.my","jobstreet.com.my","sinchew.com.my","iflix.com","chinapress.com.my","lelong.com.my","piktochart.com","binary.com","freemalaysiatoday.com","nextshark.com","uitm.edu.my","bharian.com.my","hmetro.com.my","utusan.com.my","poslaju.com.my","1govuc.gov.my","hongleongconnect.my","sinarharian.com.my","nst.com.my","astroawani.com","cari.com.my","anm.gov.my","mindvalley.com","jobstreet.com","um.edu.my","kwsp.gov.my","damasgate.com","ukm.my","thrivethemes.com","malaysiandigest.com","upm.edu.my","hasil.gov.my","orientaldaily.com.my","ptptn.gov.my","g2g.com"]



### EXCEL SPREADSHEET TOP SET-UP ###

wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row = 1, column = 1).value ="Domain"
ws.cell(row = 1, column = 2).value ="Country #1"
ws.cell(row = 1, column = 3).value ="Value"
ws.cell(row = 1, column = 4).value ="Country #2"
ws.cell(row = 1, column = 5).value ="Value"
ws.cell(row = 1, column = 6).value ="Country #3"
ws.cell(row = 1, column = 7).value ="Value"
ws.cell(row = 1, column = 8).value ="Country #4"
ws.cell(row = 1, column = 9).value ="Value"
ws.cell(row = 1, column = 10).value ="Country #5"
ws.cell(row = 1, column = 11).value ="Value"

### ACTUAL SCRAPING BEGINS HERE ###

for i in range (len(urlstring)):
    endresult1 = []
    endresult2 = []

    testurl = "https://www.alexa.com/siteinfo/" + urlstring[i]
    html = requests.get(testurl)
    soup = BeautifulSoup(html.content,'html.parser')
    geography = soup.find_all(id="demographics_div_country_table")
    lol = str(geography)
    #print (lol)


    testlist = re.finditer(r">\d+\.\d%",lol)
    for subj in testlist:
        temp = subj.group()[1:-1]
        endresult1.append(temp)

    testlist2 = re.finditer(r"alt=\"(.*?) Flag",lol)
    for subj2 in testlist2:
        temp2 = subj2.group()[5:-5]
        endresult2.append(temp2)
    
    a = len(ws["A"])+1
    ws.cell(row=len(ws["A"])+1,column = 1).value = urlstring[i]

    for x in range (len(endresult1)):
        ws.cell(row=a,column = 2*x+2).value = endresult2[x]
        ws.cell(row=a,column = 2*x+3).value = float(endresult1[x])

    wb.save("test3.xlsx")
    print("Finished analyzing domain #" + str(i+1) +".")



